# Program takes a number N from the command line and creates an NxN multiplication table in an Excel spreadsheet.
# Command line example: py multiplicationTable.py 6
# This command above should create a spreadsheet with the multiplication table all the way up to 36 on it.
# Note that this file should be named multiplicationtable.py and be place on the PATH folder.

import sys, openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, column_index_from_string

# Interprets the command line.
n = int(sys.argv[1])
# Creates the spreadsheet.
wb = openpyxl.Workbook()
sheet = wb.active
# Bold labels on first row and first column.
for i in range(1, 1+n):
    boldRow = sheet.cell(row=1, column=i).font = Font(bold=True) # bold row
    boldCol = sheet.cell(row=i, column=1).font = Font(bold=True) # bold column
    # Creating labels on first column and first row.
    labelCol = sheet['A' + str(i)].value = i
    labelRow = sheet[get_column_letter(i) + '1'].value = i
# Creates the multiplication table on the sheet from cell 'b2' and beyond.
for i in range(2, sheet.max_row + 1):
    for x in range(2, sheet.max_column + 1):
        sheet[get_column_letter(x) + str(i)].value = sheet[get_column_letter(x) + '1'].value * sheet['A' + str(i)].value

wb.save('multiplicationTable.xlsx')
print('Done.')